"""
海康监控数据整合处理系统
数据来源API名称：
1.获取编码设备列表
2.分页获取监控点资源
3.获取监控点回放取流URLv2
版本：3.0
功能：
1. 合并监控设备数据生成监控3
2. 更新videodevice.xlsx指定列数据
3. 自动备份与异常恢复
4. 数据完整性校验
"""

import os
import shutil
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
import zipfile

# ----------------- 全局配置参数 -----------------
BASE_DIR = r'D:\项目文件\南京中车园\IP点表\视频监控\监控数据对比'  # 项目根目录
BACKUP_DIR_NAME = '历史备份'  # 备份目录名称
OUTPUT_DIR_PREFIX = '结果输出'  # 结果目录前缀
TARGET_SHEET = '设备数据表'  # 目标工作表名称
COLUMN_MAPPING = {  # Excel列索引映射(从0开始)
    'device_ip': 5,  # 第6列 DeviceIP
    'replay_token': 2,  # 第3列 ReplayToken
    'third_party_key': 13,  # 第14列 ThirdPartyKey
    'playback_model': 14  # 第15列 PlaybackAccessModel
}


class DataProcessor:
    """数据处理核心类"""

    def __init__(self):
        # 初始化数据容器
        self.monitor3 = None  # 存储匹配设备数据
        self.unmatched = None  # 存储未匹配设备数据
        self.video_df = None  # 存储videodevice数据

    # ----------------- 新增关键方法 -----------------
    def extract_timestamp(self, dir_name):
        """
        从目录名解析时间戳
        :param dir_name: 目录名（格式：结果输出_YYYYMMDD_HHMMSS）
        :return: datetime对象或None
        """
        try:
            # 分割目录名各部分
            parts = dir_name.split('_')
            if len(parts) < 3:
                print(f"目录格式错误：{dir_name}（缺少日期或时间部分）")
                return None

            # 合并最后两部分为完整时间戳
            date_part = parts[-2]  # YYYYMMDD
            time_part = parts[-1]  # HHMMSS
            ts_str = f"{date_part}_{time_part}"

            return datetime.strptime(ts_str, "%Y%m%d_%H%M%S")
        except ValueError as e:
            print(f"时间戳解析失败：{dir_name} | 错误：{str(e)}")
            return None

    # ----------------- 文件校验模块 -----------------
    def validate_excel(self, path):
        """
        深度验证Excel文件完整性
        :param path: 文件路径
        :raise: 文件损坏时抛出异常
        """
        try:
            # 检查是否为有效的zip压缩包
            with zipfile.ZipFile(path) as zf:
                # 验证必须存在的文件组件
                required = ['[Content_Types].xml', 'xl/workbook.xml']
                missing = [f for f in required if f not in zf.namelist()]
                if missing:
                    raise ValueError(f"损坏文件缺失组件: {missing}")
        except zipfile.BadZipFile:
            raise ValueError("无效的Excel文件格式")

    # ----------------- 数据读取模块 -----------------
    def safe_read(self, path, cols=None):
        """
        安全读取Excel数据
        :param path: 文件路径
        :param cols: 需要读取的列名列表
        :return: 包含数据的DataFrame
        """
        # 基础文件校验
        if not os.path.isfile(path):
            raise FileNotFoundError(f"文件不存在: {os.path.basename(path)}")
        self.validate_excel(path)

        try:
            # 读取数据并验证列存在性
            df = pd.read_excel(path, usecols=cols, engine='openpyxl')
            if cols and any(c not in df.columns for c in cols):
                missing = [c for c in cols if c not in df.columns]
                raise ValueError(f"缺失必要列: {missing}")
            return df
        except Exception as e:
            raise RuntimeError(f"读取失败: {str(e)}")

    # ----------------- 数据处理模块 -----------------
    def merge_data(self):
        """合并监控1和监控2数据生成监控3"""
        # 读取原始数据
        m1 = self.safe_read(os.path.join(BASE_DIR, '监控1.xlsx'),
                            ['encodeDevIndexCode', 'cameraIndexCode'])
        m2 = self.safe_read(os.path.join(BASE_DIR, '监控2.xlsx'),
                            ['indexCode', 'ip']).rename(
            columns={'indexCode': 'encodeDevIndexCode', 'ip': 'IP'})

        # 数据清洗函数
        def clean_data(df, col):
            return (df.dropna(subset=[col])
                    .astype({col: 'str'})
                    .assign(**{col: df[col].str.strip()})
                    .drop_duplicates(subset=[col]))

        # 执行清洗
        m1_clean = clean_data(m1, 'encodeDevIndexCode')
        m2_clean = clean_data(m2, 'encodeDevIndexCode')

        # 合并数据集
        merged = pd.merge(
            m2_clean[['encodeDevIndexCode', 'IP']],
            m1_clean[['encodeDevIndexCode', 'cameraIndexCode']],
            on='encodeDevIndexCode',
            how='left',
            indicator=True
        )

        # 分离匹配/未匹配结果
        self.monitor3 = merged.query("_merge == 'both'").drop(columns='_merge')
        self.unmatched = merged.query("_merge == 'left_only'")

    # ----------------- 数据更新模块 -----------------
    def update_videodevice(self):
        """更新videodevice.xlsx文件"""
        # 获取最新监控3文件路径
        monitor3_path = self.find_latest_output()
        videodevice_path = os.path.join(BASE_DIR, 'videodevice.xlsx')

        # 创建备份文件
        backup_path = self.create_backup(videodevice_path)

        try:
            # 读取监控3映射数据
            monitor3_df = pd.read_excel(monitor3_path,
                                        usecols=['IP', 'cameraIndexCode'])
            ip_map = monitor3_df.set_index('IP')['cameraIndexCode'].to_dict()

            # 加载工作簿
            wb = load_workbook(videodevice_path)
            ws = wb.active  # 获取活动工作表

            # 遍历更新数据行
            updated_count = 0
            for row in ws.iter_rows(min_row=2):  # 跳过标题行
                # 获取设备IP
                ip = row[COLUMN_MAPPING['device_ip']].value

                # 更新ReplayToken和ThirdPartyKey
                if ip in ip_map:
                    row[COLUMN_MAPPING['replay_token']].value = ip_map[ip]
                    row[COLUMN_MAPPING['third_party_key']].value = ip_map[ip]

                # 更新PlaybackAccessModel
                if row[COLUMN_MAPPING['playback_model']].value == 1:
                    row[COLUMN_MAPPING['playback_model']].value = 3
                    updated_count += 1

            # 保存工作簿
            wb.save(videodevice_path)
            print(f"成功更新{updated_count}条记录")

            # 验证更新结果
            self.verify_update(videodevice_path, updated_count)

        except Exception as e:
            # 异常恢复处理
            print(f"错误: {str(e)}")
            self.restore_backup(backup_path, videodevice_path)
            raise
        finally:
            wb.close()

    # ----------------- 辅助功能模块 -----------------
    def find_latest_output(self):
        """查找最新监控3文件（完整实现）"""
        try:
            # 获取所有结果目录
            output_dirs = [
                d for d in os.listdir(BASE_DIR)
                if d.startswith(f"{OUTPUT_DIR_PREFIX}_")
                   and os.path.isdir(os.path.join(BASE_DIR, d))
            ]

            if not output_dirs:
                raise FileNotFoundError("未找到结果输出目录")

            # 提取有效时间戳目录
            valid_dirs = []
            for d in output_dirs:
                dt = self.extract_timestamp(d)  # 正确调用类方法
                if dt:
                    valid_dirs.append((dt, d))

            if not valid_dirs:
                raise ValueError("没有有效时间戳目录")

            # 获取最新目录
            sorted_dirs = sorted(valid_dirs, key=lambda x: x[0], reverse=True)
            latest_dir = sorted_dirs[0][1]
            monitor3_path = os.path.join(BASE_DIR, latest_dir, "监控3.xlsx")

            # 最终验证
            if not os.path.isfile(monitor3_path):
                raise FileNotFoundError(f"监控3文件不存在：{monitor3_path}")

            return monitor3_path

        except Exception as e:
            error_msg = (
                "监控3文件定位失败，请检查：\n"
                "1. 目录命名是否满足 [结果输出_日期_时间] 格式\n"
                "2. 是否已成功生成监控3.xlsx\n"
                f"技术细节：{str(e)}"
            )
            raise RuntimeError(error_msg)

    def create_backup(self, path):
        """创建带时间戳的备份文件"""
        backup_dir = os.path.join(BASE_DIR, BACKUP_DIR_NAME)
        os.makedirs(backup_dir, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = os.path.join(backup_dir, f"backup_{timestamp}.xlsx")
        shutil.copy2(path, backup_path)
        print(f"已创建备份: {backup_path}")
        return backup_path

    def verify_update(self, path, expected):
        """验证更新结果"""
        df = pd.read_excel(path)
        actual = df[df['PlaybackAccessModel'] == 3].shape[0]
        if actual != expected:
            raise ValueError(f"验证失败，预期{expected}条，实际{actual}条")

    def restore_backup(self, backup, target):
        """恢复备份文件"""
        print("正在恢复备份...")
        shutil.copy2(backup, target)
        print("备份已恢复")

    # ----------------- 报告生成模块 -----------------
    def generate_report(self):
        """生成结果报告"""
        # 创建带时间戳的输出目录
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_dir = os.path.join(BASE_DIR, f"{OUTPUT_DIR_PREFIX}_{timestamp}")
        os.makedirs(output_dir, exist_ok=True)

        # 保存监控3数据
        output_path = os.path.join(output_dir, '监控3.xlsx')
        with pd.ExcelWriter(output_path) as writer:
            self.monitor3.to_excel(writer, index=False, sheet_name='匹配设备')
            if not self.unmatched.empty:
                self.unmatched.to_excel(writer, index=False, sheet_name='未匹配设备')

        print(f"报告已生成: {output_path}")


# ----------------- 主程序入口 -----------------
def main():
    processor = DataProcessor()
    try:
        print("=== 开始处理流程 ===")

        # 执行数据合并
        processor.merge_data()

        # 生成监控3报告
        processor.generate_report()

        # 更新videodevice数据
        processor.update_videodevice()

        print("=== 处理成功 ===")
        print(f"匹配设备数: {len(processor.monitor3)}")
        print(f"未匹配设备: {len(processor.unmatched)}")

    except Exception as e:
        print(f"\n!!! 处理失败: {str(e)}")
    finally:
        print("=== 流程结束 ===")


if __name__ == "__main__":
    main()
